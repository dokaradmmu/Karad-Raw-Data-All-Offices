# ============================================
# KARAD DIVISION REPORT GENERATOR
# Complete Streamlit Application
# ============================================

import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os
import json

# ============================================
# PAGE CONFIGURATION
# ============================================
st.set_page_config(
    page_title="Karad Division Report Generator",
    page_icon="🏤",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================
# SESSION STATE INITIALIZATION
# ============================================
if 'master_uploaded' not in st.session_state:
    st.session_state.master_uploaded = False
    st.session_state.master_data = None
    st.session_state.master_filename = None
    st.session_state.all_files = {}
    st.session_state.processed = False

# ============================================
# CUSTOM CSS FOR BETTER UI
# ============================================
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #1a3a5c 0%, #2a5a8c 100%);
        padding: 2rem;
        border-radius: 12px;
        color: white;
        margin-bottom: 2rem;
    }
    .main-header h1 {
        margin: 0;
        font-size: 2.5rem;
        font-weight: 700;
    }
    .main-header p {
        margin: 0.5rem 0 0 0;
        opacity: 0.9;
        font-size: 1.1rem;
    }
    .kpi-badge {
        display: inline-block;
        padding: 0.25rem 1rem;
        border-radius: 20px;
        font-weight: 600;
        font-size: 0.85rem;
        margin: 0.25rem;
    }
    .kpi-badge-delivery { background: #e3f2fd; color: #0d47a1; }
    .kpi-badge-dss { background: #e8f5e9; color: #1b5e20; }
    .kpi-badge-cod { background: #fff3e0; color: #e65100; }
    .kpi-badge-lb { background: #fce4ec; color: #880e4f; }
    .upload-section {
        background: #f8fafc;
        padding: 1.5rem;
        border-radius: 10px;
        border: 2px dashed #c5d0de;
        margin-bottom: 1rem;
    }
    .file-status {
        padding: 0.25rem 0.75rem;
        border-radius: 12px;
        font-size: 0.8rem;
        font-weight: 600;
    }
    .file-status.uploaded { background: #e8f5e9; color: #2e7d32; }
    .file-status.pending { background: #fff3e0; color: #e65100; }
    .file-status.missing { background: #ffebee; color: #c62828; }
    .success-box {
        background: #e8f5e9;
        padding: 1rem;
        border-radius: 8px;
        border-left: 4px solid #2e7d32;
        margin: 1rem 0;
    }
    .info-box {
        background: #e3f2fd;
        padding: 1rem;
        border-radius: 8px;
        border-left: 4px solid #0d47a1;
        margin: 1rem 0;
    }
    .stButton button {
        background: #1a3a5c;
        color: white;
        font-weight: 600;
        padding: 0.5rem 2rem;
        border-radius: 8px;
        border: none;
        transition: all 0.3s;
    }
    .stButton button:hover {
        background: #0f2a44;
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(26,58,92,0.3);
    }
    .stButton button:disabled {
        opacity: 0.5;
        cursor: not-allowed;
        transform: none;
    }
    .divider {
        border-top: 2px solid #eef2f7;
        margin: 2rem 0;
    }
    .progress-text {
        font-weight: 600;
        color: #1a3a5c;
    }
</style>
""", unsafe_allow_html=True)

# ============================================
# HEADER
# ============================================
col1, col2 = st.columns([3, 1])
with col1:
    st.markdown("""
    <div class="main-header">
        <h1>🏤 Karad Division Report Generator</h1>
        <p>Upload source files to generate a fully formatted Excel report with KPIs, alerts, and actionable insights</p>
    </div>
    """, unsafe_allow_html=True)

with col2:
    st.image("https://img.icons8.com/color/96/000000/india-post.png", width=80)

# KPI Summary Badges
st.markdown("""
<div style="display: flex; gap: 12px; flex-wrap: wrap; margin-bottom: 1.5rem;">
    <span class="kpi-badge kpi-badge-delivery">📦 Delivery KPI: 90%</span>
    <span class="kpi-badge kpi-badge-dss">📱 DSS KPI: 90%</span>
    <span class="kpi-badge kpi-badge-cod">💳 COD Digital KPI: 80%</span>
    <span class="kpi-badge kpi-badge-lb">📮 LB Clearance KPI: 100%</span>
</div>
""", unsafe_allow_html=True)

# ============================================
# SIDEBAR - FILE UPLOAD
# ============================================
with st.sidebar:
    st.header("📁 File Upload")
    
    st.markdown("""
    <div class="info-box">
        <strong>📋 Instructions:</strong><br>
        1. Upload <strong>Master File</strong> first (once per session)<br>
        2. Upload <strong>all 8 data files</strong><br>
        3. Click <strong>"Generate Report"</strong>
    </div>
    """, unsafe_allow_html=True)
    
    st.divider()
    
    # Master File Upload (with replace option)
    st.subheader("📋 Master File")
    
    col_master1, col_master2 = st.columns([3, 1])
    with col_master1:
        master_file = st.file_uploader(
            "Office Master File",
            type=['xlsx', 'xls'],
            key="master_uploader",
            help="Upload the Office Master file. This is the single source of truth for office identity and hierarchy."
        )
    
    with col_master2:
        if st.session_state.master_uploaded:
            if st.button("🔄 Replace", type="secondary", use_container_width=True):
                st.session_state.master_uploaded = False
                st.session_state.master_data = None
                st.session_state.master_filename = None
                st.rerun()
    
    if master_file:
        try:
            master_df = pd.read_excel(master_file, sheet_name=0)
            master_df.columns = master_df.columns.str.strip()
            st.session_state.master_data = master_df
            st.session_state.master_filename = master_file.name
            st.session_state.master_uploaded = True
            st.success(f"✅ Master loaded: {len(master_df)} offices")
        except Exception as e:
            st.error(f"Error loading master file: {e}")
    else:
        if st.session_state.master_uploaded:
            st.success(f"✅ Master loaded: {st.session_state.master_filename}")
        else:
            st.warning("⏳ Master file not uploaded")
    
    st.divider()
    
    # Data Files Upload
    st.subheader("📊 Data Files")
    
    file_config = {
        "delRange": {"label": "📦 Delivery (Range)", "type": "csv", "required": True},
        "delSingle": {"label": "📦 Delivery (Latest)", "type": "csv", "required": True},
        "dssRange": {"label": "📱 DSS (Range)", "type": "csv", "required": True},
        "dssSingle": {"label": "📱 DSS (Latest)", "type": "csv", "required": True},
        "codRange": {"label": "💳 COD Digital (Range)", "type": "csv", "required": True},
        "codSingle": {"label": "💳 COD Digital (Latest)", "type": "csv", "required": True},
        "lbRange": {"label": "📮 LB Clearance (Range)", "type": "csv", "required": True},
        "lbSingle": {"label": "📮 LB Clearance (Latest)", "type": "csv", "required": True},
    }
    
    uploaded_files = {}
    all_uploaded = True
    
    for key, config in file_config.items():
        col1, col2 = st.columns([3, 1])
        with col1:
            file = st.file_uploader(
                config["label"],
                type=config["type"],
                key=f"file_{key}",
                help=f"Expected: {key.replace('Range', '01-11').replace('Single', '11')}.csv"
            )
            if file:
                uploaded_files[key] = file
                st.session_state.all_files[key] = file
        with col2:
            if file:
                st.markdown(f'<span class="file-status uploaded">✅ Uploaded</span>', unsafe_allow_html=True)
            elif key in st.session_state.all_files:
                st.markdown(f'<span class="file-status uploaded">✅ Uploaded</span>', unsafe_allow_html=True)
            else:
                st.markdown(f'<span class="file-status pending">⏳ Pending</span>', unsafe_allow_html=True)
                all_uploaded = False
    
    # Store uploaded files in session state
    st.session_state.all_files.update(uploaded_files)
    
    st.divider()
    
    # File Status Summary
    st.subheader("📋 Upload Status")
    
    total_files = 9  # 8 data + 1 master
    uploaded_count = 0
    
    if st.session_state.master_uploaded:
        uploaded_count += 1
        st.markdown("✅ Master file: **Uploaded**")
    else:
        st.markdown("❌ Master file: **Missing**")
    
    for key in file_config.keys():
        if key in st.session_state.all_files:
            uploaded_count += 1
            st.markdown(f"✅ {file_config[key]['label']}: **Uploaded**")
        else:
            st.markdown(f"❌ {file_config[key]['label']}: **Missing**")
    
    st.caption(f"**{uploaded_count}/{total_files}** files uploaded")
    
    all_files_ready = (uploaded_count == total_files)

# ============================================
# DATA PROCESSING FUNCTIONS
# ============================================
@st.cache_data
def parse_csv(file):
    """Parse CSV file to DataFrame"""
    try:
        df = pd.read_csv(file)
        df.columns = df.columns.str.strip()
        return df
    except Exception as e:
        st.error(f"Error parsing {file.name}: {e}")
        return None

def get_office_data(df, office_id, column):
    """Get value from dataframe for specific office"""
    if df is None or df.empty:
        return None
    id_cols = ['office-id', 'office_id', 'Office ID']
    for id_col in id_cols:
        if id_col in df.columns:
            row = df[df[id_col] == office_id]
            if not row.empty and column in df.columns:
                val = row[column].iloc[0]
                if pd.isna(val):
                    return None
                return val
    return None

def get_cod_data(df, office_id):
    """Get aggregated COD data for an office"""
    if df is None or df.empty:
        return None
    
    id_cols = ['Office ID', 'office_id', 'office-id']
    for id_col in id_cols:
        if id_col in df.columns:
            filtered = df[df[id_col] == office_id]
            if not filtered.empty:
                total_col = 'Total COD Count'
                digital_col = 'COD Digital Count'
                total = filtered[total_col].sum() if total_col in df.columns else 0
                digital = filtered[digital_col].sum() if digital_col in df.columns else 0
                return {'totalCount': total, 'digitalCount': digital}
    return None

def calculate_kpi(row_data, numerator, denominator):
    """Calculate KPI percentage"""
    if row_data is None:
        return None
    num = row_data.get(numerator, 0) if isinstance(row_data, dict) else row_data[numerator] if numerator in row_data else 0
    den = row_data.get(denominator, 0) if isinstance(row_data, dict) else row_data[denominator] if denominator in row_data else 0
    if pd.isna(num) or pd.isna(den) or den == 0:
        return None
    return (num / den) * 100

# ============================================
# EXCEL GENERATION WITH FORMATTING
# ============================================
def generate_excel_report(master_df, files):
    """Generate fully formatted Excel workbook"""
    
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Sort master data
    master_sorted = master_df.sort_values(
        ['Sub Division Name', 'Sub Office Name', 'Office Type Code']
    ).reset_index(drop=True)
    
    # Create Sheet 1: Raw Data
    ws1 = wb.create_sheet("Raw Data", 0)
    create_raw_data_sheet(ws1, master_sorted, files)
    
    # Create Sheet 2: Sub Division Summary
    ws2 = wb.create_sheet("Sub Division Summary", 1)
    create_summary_sheet(ws2, master_sorted, files)
    
    # Create Sheet 3: Sub Office Details
    ws3 = wb.create_sheet("Sub Office Details", 2)
    create_detail_sheet(ws3, master_sorted, files)
    
    return wb

def create_raw_data_sheet(ws, master_df, files):
    """Create Raw Data sheet with full formatting"""
    
    # Title
    ws.merge_cells('A1:AD1')
    title_cell = ws.cell(row=1, column=1, value="KARAD DIVISION — CONSOLIDATED PERFORMANCE & DEFAULTER REPORT")
    title_cell.font = Font(bold=True, size=14, color="1a3a5c")
    title_cell.alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:AD2')
    subtitle_cell = ws.cell(row=2, column=1, value="Sheet 1: Raw Data (All 288 Offices)  |  Reporting Period: Based on uploaded files")
    subtitle_cell.font = Font(size=10, italic=True, color="5a6a7a")
    subtitle_cell.alignment = Alignment(horizontal="center")
    
    # Headers
    headers = [
        "Sr. No.", "Sub Division", "Sub Office", "Office Name", "Office Type",
        "Postman", "Invoice", "Deposit", "Delivery % (Period)",
        "Postman", "Invoice", "Deposit", "Delivery % (Latest)",
        "PDM Articles", "DSS Articles", "DSS % (Period)",
        "PDM Articles", "DSS Articles", "DSS % (Latest)",
        "COD Total", "COD Digital", "COD Digital % (Period)",
        "COD Total", "COD Digital", "COD Digital % (Latest)",
        "Total LB", "LB Clearance % (Period)", "LB Clearance % (Latest)",
        "Office ID", "Universe"
    ]
    
    # Header formatting
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1a3a5c", end_color="1a3a5c", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Light blue sub-headers for sections
    section_cols = {
        6: "DELIVERY — PERIOD",
        10: "DELIVERY — LATEST",
        14: "DSS — PERIOD",
        17: "DSS — LATEST",
        20: "COD — PERIOD",
        23: "COD — LATEST",
        26: "LETTER BOX"
    }
    
    subheader_fill = PatternFill(start_color="d4e0f0", end_color="d4e0f0", fill_type="solid")
    subheader_font = Font(bold=True, size=9, color="1a3a5c")
    
    for col, label in section_cols.items():
        cell = ws.cell(row=4, column=col, value=label)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    row_num = 5
    for idx, (_, office) in enumerate(master_df.iterrows(), 1):
        office_id = office['Office ID']
        office_type = office['Office Type Code']
        universe = 'B' if office_type == 'BPO' else 'A'
        
        # Fetch data
        del_range = get_office_data(files.get('delRange'), office_id, 'invoice-count')
        del_single = get_office_data(files.get('delSingle'), office_id, 'invoice-count')
        
        row = [
            idx,
            office.get('Sub Division Name', ''),
            office.get('Sub Office Name', ''),
            office.get('Office Name', ''),
            office_type,
        ]
        
        # Delivery Period
        row.append(get_office_data(files.get('delRange'), office_id, 'postman-count') or '')
        row.append(get_office_data(files.get('delRange'), office_id, 'invoice-count') or '')
        row.append(get_office_data(files.get('delRange'), office_id, 'deposit-count') or '')
        
        del_row = get_office_data(files.get('delRange'), office_id, None)
        del_pct = calculate_kpi(del_row, 'invoice-count', 'deposit-count') if isinstance(del_row, dict) else None
        row.append(del_pct if del_pct is not None else '')
        
        # Delivery Single
        row.append(get_office_data(files.get('delSingle'), office_id, 'postman-count') or '')
        row.append(get_office_data(files.get('delSingle'), office_id, 'invoice-count') or '')
        row.append(get_office_data(files.get('delSingle'), office_id, 'deposit-count') or '')
        
        del_row_single = get_office_data(files.get('delSingle'), office_id, None)
        del_pct_single = calculate_kpi(del_row_single, 'invoice-count', 'deposit-count') if isinstance(del_row_single, dict) else None
        row.append(del_pct_single if del_pct_single is not None else '')
        
        # DSS Period
        row.append(get_office_data(files.get('dssRange'), office_id, 'total_pdm_art_count') or '')
        row.append(get_office_data(files.get('dssRange'), office_id, 'total_dss_art_count') or '')
        
        dss_row = get_office_data(files.get('dssRange'), office_id, None)
        dss_pct = calculate_kpi(dss_row, 'total_dss_art_count', 'total_pdm_art_count') if isinstance(dss_row, dict) else None
        row.append(dss_pct if dss_pct is not None else '')
        
        # DSS Single
        row.append(get_office_data(files.get('dssSingle'), office_id, 'total_pdm_art_count') or '')
        row.append(get_office_data(files.get('dssSingle'), office_id, 'total_dss_art_count') or '')
        
        dss_row_single = get_office_data(files.get('dssSingle'), office_id, None)
        dss_pct_single = calculate_kpi(dss_row_single, 'total_dss_art_count', 'total_pdm_art_count') if isinstance(dss_row_single, dict) else None
        row.append(dss_pct_single if dss_pct_single is not None else '')
        
        # COD Period
        cod_range = get_cod_data(files.get('codRange'), office_id)
        row.append(cod_range.get('totalCount', '') if cod_range else '')
        row.append(cod_range.get('digitalCount', '') if cod_range else '')
        cod_pct = (cod_range['digitalCount'] / cod_range['totalCount'] * 100) if cod_range and cod_range['totalCount'] > 0 else ''
        row.append(cod_pct if cod_pct != '' else '')
        
        # COD Single
        cod_single = get_cod_data(files.get('codSingle'), office_id)
        row.append(cod_single.get('totalCount', '') if cod_single else '')
        row.append(cod_single.get('digitalCount', '') if cod_single else '')
        cod_pct_single = (cod_single['digitalCount'] / cod_single['totalCount'] * 100) if cod_single and cod_single['totalCount'] > 0 else ''
        row.append(cod_pct_single if cod_pct_single != '' else '')
        
        # LB
        row.append(get_office_data(files.get('lbRange'), office_id, 'total-letterboxes') or '')
        row.append(get_office_data(files.get('lbRange'), office_id, 'clearance-percentage') or '')
        row.append(get_office_data(files.get('lbSingle'), office_id, 'clearance-percentage') or '')
        
        # Helpers
        row.append(office_id)
        row.append(universe)
        
        # Write row with alternating colors
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")
        
        row_num += 1
    
    # Column widths
    col_widths = [8, 22, 22, 25, 10, 10, 10, 10, 16, 10, 10, 10, 16, 12, 12, 16, 12, 12, 16, 12, 12, 16, 12, 12, 16, 12, 16, 16, 12, 10]
    for i, width in enumerate(col_widths, 1):
        col_letter = chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)
        ws.column_dimensions[col_letter].width = width
    
    # Freeze panes
    ws.freeze_panes = 'A6'
    
    # Add footer note
    ws.merge_cells('A30:AD30')
    footer_cell = ws.cell(row=row_num + 2, column=1, value="Note: All percentages are volume-weighted. Blank cells indicate no data available.")
    footer_cell.font = Font(size=9, italic=True, color="6a7a8a")

def create_summary_sheet(ws, master_df, files):
    """Create Sub Division Summary sheet with formatting"""
    # This will be implemented in the next step
    pass

def create_detail_sheet(ws, master_df, files):
    """Create Sub Office Details sheet with formatting"""
    # This will be implemented in the next step
    pass

# ============================================
# MAIN APP LOGIC
# ============================================
st.divider()

# Main area - show status and generate button
col_status, col_button = st.columns([3, 1])

with col_status:
    if all_files_ready:
        st.markdown("""
        <div class="success-box">
            ✅ <strong>All files uploaded successfully!</strong><br>
            Ready to generate the formatted Excel report.
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown("""
        <div class="info-box">
            📥 <strong>Upload all required files</strong><br>
            Please upload the Master file and all 8 data files to enable report generation.
        </div>
        """, unsafe_allow_html=True)

with col_button:
    generate_button = st.button(
        "🚀 Generate Report",
        type="primary",
        disabled=not all_files_ready,
        use_container_width=True
    )

# Report generation logic
if generate_button and all_files_ready:
    try:
        with st.spinner("🔄 Processing files and generating formatted report..."):
            # Parse data files
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            status_text.text("Loading data files...")
            files_data = {}
            for key in ['delRange', 'delSingle', 'dssRange', 'dssSingle', 
                       'codRange', 'codSingle', 'lbRange', 'lbSingle']:
                if key in st.session_state.all_files:
                    files_data[key] = parse_csv(st.session_state.all_files[key])
            
            progress_bar.progress(50)
            
            status_text.text("Generating Excel workbook with formatting...")
            wb = generate_excel_report(st.session_state.master_data, files_data)
            
            progress_bar.progress(80)
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            progress_bar.progress(100)
            status_text.text("✅ Report ready for download!")
            
            # Download button
            filename = f"Karad_Division_Report_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
            st.download_button(
                label="📥 Download Formatted Report",
                data=output,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )
            
            st.success("✅ Report generated successfully! Click the download button above.")
            
    except Exception as e:
        st.error(f"❌ Error generating report: {str(e)}")
        import traceback
        st.code(traceback.format_exc())

# ============================================
# FOOTER
# ============================================
st.divider()
st.caption(f"🏤 Karad Division Report Generator v1.0 | Last updated: {datetime.now().strftime('%d-%m-%Y %H:%M')}")
